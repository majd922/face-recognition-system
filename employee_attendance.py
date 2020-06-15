import asyncio
import io
import glob
import os
import sys
import time
from datetime import datetime
import cv2
import uuid
import requests
from urllib.parse import urlparse
from io import BytesIO
from PIL import Image, ImageDraw
from azure.cognitiveservices.vision.face import FaceClient
from msrest.authentication import CognitiveServicesCredentials
from azure.cognitiveservices.vision.face.models import TrainingStatusType, Person, SnapshotObjectType, \
    OperationStatusType

# Set the FACE_SUBSCRIPTION_KEY environment variable with your key as the value.
# This key will serve all examples in this document.
KEY = os.environ['FACE_SUBSCRIPTION_KEY']

# Set the FACE_ENDPOINT environment variable with the endpoint from your Face service in Azure.
# This endpoint will be used in all examples in this quickstart.
ENDPOINT = os.environ['FACE_ENDPOINT']

# Create an authenticated FaceClient.
face_client = FaceClient(ENDPOINT, CognitiveServicesCredentials(KEY))

# Used in the Person Group Operations,  Snapshot Operations, and Delete Person Group examples.
# You can call list_person_groups to print a list of preexisting PersonGroups.
# SOURCE_PERSON_GROUP_ID should be all lowercase and alphanumeric. For example, 'mygroupname' (dashes are OK).
PERSON_GROUP_ID = 'my-unique-person-group'

# Used for the Snapshot and Delete Person Group examples.
TARGET_PERSON_GROUP_ID = str(uuid.uuid4())  # assign a random ID (or name it anything)
'''
 Identify a face against a defined PersonGroup
'''
cam = cv2.VideoCapture(0)

while (True):

    # reading from frame
    ret, frame = cam.read()

    if ret:
        cv2.imshow("preview", frame)
        k = cv2.waitKey(1)

        if k % 256 == 27:
            # ESC pressed
            print("Escape hit, closing...")
            break
        elif k % 256 == 32:
            # SPACE pressed
            name = 'test employee.jpg'
            print('Creating...' + name)
            cv2.imwrite(name, frame)
            print("{} written!".format(name))
            break
    else:
        break

# Release all space and windows once done
cam.release()
cv2.destroyAllWindows()

# Group image for testing against
group_photo = str(name)
IMAGES_FOLDER = os.path.join(os.path.dirname(os.path.realpath(__file__)))
# Get test image
test_image_array = glob.glob(os.path.join(IMAGES_FOLDER, group_photo))
image = open(test_image_array[0], 'r+b')

# Detect faces
face_ids = []
faces = face_client.face.detect_with_stream(image)
for face in faces:
    face_ids.append(face.face_id)

# Identify faces
results = face_client.face.identify(face_ids, PERSON_GROUP_ID)
print ('Identifying faces in {}'.format(PERSON_GROUP_ID))
now = datetime.now()
for person in results:
        if person.candidates:
             nameOfFound = "Unknown"
             IDs = [guy for guy in face_client.person_group_person.list(person_group_id=PERSON_GROUP_ID)]
             for i in IDs:
                 if i.person_id == person.candidates[0].person_id:
                               nameOfFound = i.name
             print ('Person for face ID: {} \nis identified in: {} \nwith a confidence of: {}. \nEmployee name is:  {}'.format(person.face_id, os.path.basename(image.name), person.candidates[0].confidence, nameOfFound),"\ndate and time login: ",now.strftime("%d/%m/%Y %H:%M:%S"),
                    '\nThe results are saved in "Employees Attendance.xlsx" file ')


from openpyxl import Workbook
from openpyxl import load_workbook
import pathlib
file = pathlib.Path('Employees Attendance.xlsx')
if file.exists ():
    wb = load_workbook("Employees Attendance.xlsx")
    # Select First Worksheet
    ws = wb.worksheets[0]

    rows = (
        (nameOfFound, now.strftime("%d/%m/%Y %H:%M:%S"), person.candidates[0].confidence),
    )

    for row in rows:
        ws.append(row)

    wb.save('Employees Attendance.xlsx')

else:
    book = Workbook()
    sheet = book.active
    sheet['A1'] = "Name"
    sheet['B1'] = "Date & Time"
    sheet['C1'] = "Confidence"

    rows = (
        (nameOfFound, now.strftime("%d/%m/%Y %H:%M:%S"), person.candidates[0].confidence),
    )

    for row in rows:
        sheet.append(row)

    book.save('Employees Attendance.xlsx')
####
